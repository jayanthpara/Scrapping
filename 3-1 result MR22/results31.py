import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from webdriver_manager.chrome import ChromeDriverManager

# Set up the Chrome driver
options = webdriver.ChromeOptions()
options.add_argument('--headless')
options.add_argument('--no-sandbox')
options.add_argument('--disable-dev-shm-usage')
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
driver.set_page_load_timeout(10)

url = "https://mrecresults.mrecexams.com/StudentResult/Index?Id=615&ex76brs22fbmm=6g6yTaG6vFSsHkLfnS"

# Roll number generator
def generate_roll_numbers():
    base = "23j45a67"
    for i in range(1, 22):
        yield f"{base}{i:02}"
    #for prefix in 'abcdefghjklmnopqr':
        #for i in range(0, 10):
            #yield f"{base}{prefix}{i}"

# Excel setup
wb = Workbook()
ws = wb.active
ws.append([
    "S.No", "Roll Number", "Name", "CGPA", "SGPA", "Percentage",
    "Matrices and Calculus", "Engineering Chemistry", "Applied Physics",
    "Programming for Problem Solving", "Engineering Drawing", "Backlogs"
])

red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

# Main data extractor
def get_data(roll, sno):
    try:
        driver.get(url)

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="myApp"]/body/div[2]/form/div/div/div[1]/div/input[1]'))).send_keys(roll)
        driver.find_element(By.XPATH, '//*[@id="myApp"]/body/div[2]/form/div/div/div[2]/div/input').click()

        time.sleep(4)  # Wait for result to load

        name = driver.find_element(By.XPATH, '//*[@id="myApp"]/body/div[2]/div[2]/div/div[1]/span[4]').text
        cgpa = driver.find_element(By.ID, f"cgpa_{roll.upper()}").text
        sgpa = driver.find_element(By.ID, f"sgpa_{roll.upper()}").text
        percentage = driver.find_element(By.XPATH, '//*[@id="myApp"]/body/div[2]/div[2]/div/div[3]/div/table[2]/tbody/tr[13]/td[1]').text.replace('%', '').strip()

        marks = {
            "matrices": driver.find_element(By.XPATH, '//*[@id="myApp"]/body/div[2]/div[2]/div/div[3]/div/table[2]/tbody/tr[4]/td[7]').text,
            "chemistry": driver.find_element(By.XPATH, '//*[@id="myApp"]/body/div[2]/div[2]/div/div[3]/div/table[2]/tbody/tr[5]/td[7]').text,
            "physics": driver.find_element(By.XPATH, '//*[@id="myApp"]/body/div[2]/div[2]/div/div[3]/div/table[2]/tbody/tr[6]/td[7]').text,
            "programming": driver.find_element(By.XPATH, '//*[@id="myApp"]/body/div[2]/div[2]/div/div[3]/div/table[2]/tbody/tr[7]/td[7]').text,
            "drawing": driver.find_element(By.XPATH, '//*[@id="myApp"]/body/div[2]/div[2]/div/div[3]/div/table[2]/tbody/tr[8]/td[7]').text
        }

        backlog_count = sum(1 for mark in marks.values() if mark.isdigit() and int(mark) < 21)

        row = [
            sno, roll.upper(), name, cgpa, sgpa, percentage,
            marks['matrices'], marks['chemistry'], marks['physics'],
            marks['programming'], marks['drawing'], backlog_count
        ]
        ws.append(row)

        for i, mark in enumerate([marks['matrices'], marks['chemistry'], marks['physics'], marks['programming'], marks['drawing']]):
            if mark.isdigit() and int(mark) < 21:
                cell = ws.cell(row=sno + 1, column=7 + i)
                cell.fill = red_fill

    except (TimeoutException, NoSuchElementException) as e:
        print(f"Skipping {roll.upper()}: {e}")

# Run script
sno = 1
for roll in generate_roll_numbers():
    print(f"Processing {roll.upper()}...")
    get_data(roll, sno)
    sno += 1

# Save the Excel file
wb.save("MREC_Results.xlsx")
driver.quit()
print("Scraping completed and Excel file saved as MREC_Results.xlsx")
